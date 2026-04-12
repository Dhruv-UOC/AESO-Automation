"""
utils/bus_listing.py
---------------------
Standalone utility that opens a PSS/E .sav file and exports a complete
bus listing to Excel and the console.

Run this ONCE after you have your .sav file but before running any study.
Use the output to fill in:
  • Bus_Numbers sheet in study_scope_data.xlsx
  • Source Bus Number, POI Bus Number, TS Fault Bus Number in Project_Info sheet
  • From Bus No / To Bus No columns in TS_Contingencies and PV_Contingencies sheets
  • Bus No column in SC_Substations sheet

Output
------
  • Console table: all buses sorted by base kV descending
  • Excel file:    output/results/bus_listing_<timestamp>.xlsx
    Sheets:
      All_Buses      — complete bus list (number, name, base kV, type, voltage)
      Generators     — type 2 and 3 buses (generator and swing buses)
      High_Voltage   — 138 kV and above (for SC and PV monitoring)
      Substations    — unique substation names extracted from bus names

Usage
-----
    # From project root (all settings from config/settings.py):
    python utils/bus_listing.py

    # With explicit paths:
    python utils/bus_listing.py --sav D:\\Final_Project\\files\\Projects\\P2611\\Cases\\2029SW_SP_HR.sav
                                --output D:\\...\\output\\results

    # Filter by substation name (partial match, case-insensitive):
    python utils/bus_listing.py --filter lanfine

    # Show only 240 kV buses:
    python utils/bus_listing.py --kv 240
"""

import argparse
import logging
import os
import sys
from datetime import datetime
from typing import List, Optional

# ── Ensure project root is on path ───────────────────────────────────────────
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── USER CONFIGURATION ────────────────────────────────────────────────────────
# Set your .sav file path here — this will be used if no --sav flag is provided
USER_SAV_FILE = r"D:\Final_Project\files\Projects\P1828\cases\2021RC_2025SP_South-High_Tie-Econ_Solar-0.95_p.sav"
# Example alternatives:
#   r"D:\Final_Project\files\Projects\P1828\cases\2021RC_2025SP_South-High_Tie-Econ_Solar-0.95_p.sav"
#   r"D:\Final_Project\files\cases\sample.sav"


# ── Data container ────────────────────────────────────────────────────────────

class BusEntry:
    """One bus from the PSS/E model."""
    __slots__ = (
        "bus_number", "bus_name", "base_kv",
        "bus_type", "type_label",
        "voltage_pu", "voltage_kv", "angle_deg",
        "area_no",
    )

    BUS_TYPE_LABELS = {
        1: "Load",
        2: "Generator",
        3: "Slack/Swing",
        4: "Isolated",
    }

    def __init__(
        self,
        bus_number: int,
        bus_name:   str,
        base_kv:    float,
        bus_type:   int,
        voltage_pu: float,
        angle_deg:  float,
        area_no:    int,
    ):
        self.bus_number  = bus_number
        self.bus_name    = bus_name.strip()
        self.base_kv     = base_kv
        self.bus_type    = bus_type
        self.type_label  = self.BUS_TYPE_LABELS.get(bus_type, f"Type {bus_type}")
        self.voltage_pu  = round(voltage_pu, 5)
        self.voltage_kv  = round(voltage_pu * base_kv, 3)
        self.angle_deg   = round(angle_deg, 4)
        self.area_no     = area_no


# ── Bus listing class ─────────────────────────────────────────────────────────

class BusListing:
    """
    Extracts and exports bus data from a PSS/E .sav file.

    Parameters
    ----------
    psse_path : str
        Path to PSS/E PSSBIN directory.
    psse_version : int
        PSS/E major version number.
    """

    def __init__(self, psse_path: str, psse_version: int = 35):
        self.psse_path    = psse_path
        self.psse_version = psse_version
        self._psspy       = None
        self.buses: List[BusEntry] = []

    # ── Public ────────────────────────────────────────────────────────────────

    def run(
        self,
        sav_path:   str,
        output_dir: str,
        name_filter:Optional[str] = None,
        kv_filter:  Optional[float] = None,
    ) -> str:
        """
        Load the .sav file, extract all buses, print table, save Excel.

        Parameters
        ----------
        sav_path : str
            Path to the .sav file.
        output_dir : str
            Directory where the Excel file is saved.
        name_filter : str, optional
            Only show buses whose name contains this string (case-insensitive).
        kv_filter : float, optional
            Only show buses at this base kV level.

        Returns
        -------
        str : Path to the saved Excel file.
        """
        self._init_psse()
        self._load_case(sav_path)
        self._solve_power_flow()
        self.buses = self._extract_buses()

        if not self.buses:
            logger.error("No buses extracted from %s.", sav_path)
            return ""

        logger.info("Extracted %d buses from: %s", len(self.buses), sav_path)

        # Apply filters for console display
        display_buses = self.buses
        if name_filter:
            display_buses = [
                b for b in display_buses
                if name_filter.lower() in b.bus_name.lower()
            ]
            logger.info(
                "Filter '%s': %d buses matched.", name_filter, len(display_buses)
            )
        if kv_filter:
            display_buses = [
                b for b in display_buses
                if abs(b.base_kv - kv_filter) < 1.0
            ]
            logger.info(
                "Filter %.0f kV: %d buses matched.", kv_filter, len(display_buses)
            )

        self._print_table(display_buses)

        # Save Excel with ALL buses (not filtered)
        excel_path = self._save_excel(output_dir)
        return excel_path

    # ── Private: PSS/E calls ──────────────────────────────────────────────────

    def _init_psse(self) -> None:
        """Add PSS/E to sys.path and import psspy."""
        import sys as _sys

        psspy37 = os.path.abspath(
            os.path.join(self.psse_path, "..", "PSSPY37")
        )
        for p in (psspy37, self.psse_path):
            if os.path.isdir(p) and p not in _sys.path:
                _sys.path.insert(0, p)

        os.environ["PATH"] = (
            self.psse_path + os.pathsep + os.environ.get("PATH", "")
        )

        try:
            import psspy  # type: ignore
            self._psspy = psspy
        except ImportError as exc:
            raise RuntimeError(
                f"Could not import psspy.\n"
                f"  PSSBIN  : {self.psse_path}\n"
                f"  PSSPY37 : {psspy37}\n"
                f"  Verify PSS/E is installed and the licence is active.\n"
                f"  Reason  : {exc}"
            ) from exc

        ret = self._psspy.psseinit(150000)
        if ret != 0:
            raise RuntimeError(
                f"psspy.psseinit() returned error code {ret}. "
                f"Check PSS/E licence."
            )
        # Suppress PSS/E output
        self._psspy.report_output(6, "", [])
        self._psspy.progress_output(6, "", [])
        logger.info("PSS/E %d initialised.", self.psse_version)

    def _load_case(self, sav_path: str) -> None:
        """Load the .sav case file."""
        sav_path = os.path.abspath(sav_path)
        if not os.path.isfile(sav_path):
            raise FileNotFoundError(
                f"SAV file not found: {sav_path}"
            )
        ret = self._psspy.case(sav_path)
        if ret != 0:
            raise RuntimeError(
                f"psspy.case() returned error code {ret} for: {sav_path}"
            )
        logger.info("Loaded case: %s", sav_path)

    def _solve_power_flow(self) -> None:
        """Solve power flow to get bus voltages."""
        ret = self._psspy.fnsl([1, 0, 1, 1, 1, 0, 0, 0])
        if ret != 0:
            logger.warning(
                "Power flow did not converge (ret=%d). "
                "Bus voltages may show flat-start values.", ret
            )
        else:
            logger.info("Power flow solved.")

    def _extract_buses(self) -> List[BusEntry]:
        """Extract bus data arrays from PSS/E."""
        psspy = self._psspy
        buses = []

        try:
            ierr, (bus_nums,)  = psspy.abusint(-1, 1, ["NUMBER"])
            ierr, (bus_types,) = psspy.abusint(-1, 1, ["TYPE"])
            ierr, (area_nos,)  = psspy.abusint(-1, 1, ["AREA"])
            ierr, (bus_names,) = psspy.abuschar(-1, 1, ["NAME"])
            ierr, (base_kvs,)  = psspy.abusreal(-1, 1, ["BASE"])
            ierr, (voltages,)  = psspy.abusreal(-1, 1, ["PU"])
            ierr, (angles,)    = psspy.abusreal(-1, 1, ["ANGLED"])

        except Exception as exc:
            logger.error("Failed to extract bus arrays: %s", exc)
            return []

        for i, bnum in enumerate(bus_nums):
            try:
                buses.append(BusEntry(
                    bus_number = bnum,
                    bus_name   = bus_names[i] if bus_names else f"BUS_{bnum}",
                    base_kv    = base_kvs[i]  if base_kvs  else 0.0,
                    bus_type   = bus_types[i]  if bus_types  else 1,
                    voltage_pu = voltages[i]   if voltages   else 1.0,
                    angle_deg  = angles[i]     if angles     else 0.0,
                    area_no    = area_nos[i]   if area_nos   else 0,
                ))
            except (IndexError, TypeError) as exc:
                logger.debug("Skipping bus %d: %s", bnum, exc)

        # Sort by base kV descending, then bus number ascending
        buses.sort(key=lambda b: (-b.base_kv, b.bus_number))
        return buses

    # ── Private: output ───────────────────────────────────────────────────────

    def _print_table(self, buses: List[BusEntry]) -> None:
        """Print a formatted bus table to console."""
        if not buses:
            print("No buses to display.")
            return

        header = (
            f"{'Bus No':>8}  "
            f"{'Bus Name':<30}  "
            f"{'Base kV':>8}  "
            f"{'Type':<12}  "
            f"{'V (pu)':>8}  "
            f"{'V (kV)':>9}  "
            f"{'Angle':>8}  "
            f"{'Area':>5}"
        )
        sep = "-" * len(header)
        print(f"\n{sep}")
        print(f"  Bus Listing  ({len(buses)} buses)")
        print(sep)
        print(header)
        print(sep)

        for b in buses:
            print(
                f"{b.bus_number:>8}  "
                f"{b.bus_name:<30}  "
                f"{b.base_kv:>8.1f}  "
                f"{b.type_label:<12}  "
                f"{b.voltage_pu:>8.5f}  "
                f"{b.voltage_kv:>9.3f}  "
                f"{b.angle_deg:>8.4f}  "
                f"{b.area_no:>5}"
            )

        print(sep)
        print(f"  Total: {len(buses)} buses displayed.")
        print(sep)

    def _save_excel(self, output_dir: str) -> str:
        """Save full bus listing to Excel with multiple useful sheets."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
        except ImportError:
            logger.warning(
                "openpyxl not installed. Excel output skipped. "
                "Install with: pip install openpyxl"
            )
            return ""

        os.makedirs(output_dir, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(output_dir, f"bus_listing_{ts}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Styles
        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="003865")
        body_font = Font(name="Arial", size=10)
        alt_fill  = PatternFill("solid", fgColor="D6E4F0")
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")
        side      = Side(style="thin", color="AAAAAA")
        border    = Border(left=side, right=side, top=side, bottom=side)

        def add_sheet(title: str, buses: List[BusEntry]) -> None:
            ws = wb.create_sheet(title)
            ws.sheet_view.showGridLines = False

            cols = [
                ("Bus Number", 12), ("Bus Name", 32), ("Base kV", 10),
                ("Bus Type", 14),   ("V (pu)",   10), ("V (kV)",  11),
                ("Angle (deg)", 12),("Area No",  10),
            ]
            for col_idx, (col_name, width) in enumerate(cols, start=1):
                c = ws.cell(row=1, column=col_idx, value=col_name)
                c.font      = hdr_font
                c.fill      = hdr_fill
                c.alignment = center
                c.border    = border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = width

            ws.row_dimensions[1].height = 20

            for row_idx, b in enumerate(buses, start=2):
                alt  = (row_idx % 2 == 0)
                row_data = [
                    b.bus_number, b.bus_name, b.base_kv,
                    b.type_label, b.voltage_pu, b.voltage_kv,
                    b.angle_deg, b.area_no,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=value)
                    c.font      = body_font
                    c.alignment = left
                    c.border    = border
                    if alt:
                        c.fill = alt_fill

            ws.freeze_panes = "A2"

        # Sheet 1: All buses
        add_sheet("All_Buses", self.buses)

        # Sheet 2: Generators (type 2 = generator, type 3 = slack)
        generators = [b for b in self.buses if b.bus_type in (2, 3)]
        add_sheet("Generators", generators)

        # Sheet 3: High voltage (138 kV and above)
        hv_buses = [b for b in self.buses if b.base_kv >= 138.0]
        add_sheet("High_Voltage_138kV_plus", hv_buses)

        # Sheet 4: 240 kV buses
        buses_240 = [b for b in self.buses if abs(b.base_kv - 240.0) < 5.0]
        add_sheet("Buses_240kV", buses_240)

        # Sheet 5: Substations — unique substation names
        # PSS/E bus names often follow pattern "SUBSTATIONNAME 999S"
        # Extract unique substation root names
        substation_names = self._extract_substation_names()
        ws_subs = wb.create_sheet("Substation_Names")
        ws_subs.sheet_view.showGridLines = False
        ws_subs.column_dimensions["A"].width = 35
        ws_subs.column_dimensions["B"].width = 14
        ws_subs.column_dimensions["C"].width = 12
        ws_subs.column_dimensions["D"].width = 14

        for col_idx, header_text in enumerate(
            ["Substation Name (from PSS/E)", "Bus Number", "Base kV", "Bus Type"],
            start=1
        ):
            c = ws_subs.cell(row=1, column=col_idx, value=header_text)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = center
            c.border    = border

        for row_idx, (name, bus) in enumerate(substation_names, start=2):
            alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(
                [name, bus.bus_number, bus.base_kv, bus.type_label], start=1
            ):
                c = ws_subs.cell(row=row_idx, column=col_idx, value=value)
                c.font = body_font
                c.alignment = left
                c.border = border
                if alt:
                    c.fill = alt_fill
        ws_subs.freeze_panes = "A2"

        # Sheet 6: Instructions
        ws_inst = wb.create_sheet("How_To_Use", 0)
        ws_inst.sheet_view.showGridLines = False
        ws_inst.column_dimensions["A"].width = 3
        ws_inst.column_dimensions["B"].width = 60

        instructions = [
            ("HOW TO USE THIS BUS LISTING", True),
            ("", False),
            ("Step 1:", False),
            ("  Open the Bus_Numbers sheet in your study_scope_data.xlsx", False),
            ("  Copy the Bus Number and Base kV for each substation listed there", False),
            ("  from the All_Buses or Substation_Names sheet in this file.", False),
            ("", False),
            ("Step 2:", False),
            ("  In study_scope_data.xlsx Project_Info sheet, fill in:", False),
            ("    Source Bus Number  — the solar plant generator bus (Type=Generator)", False),
            ("    POI Bus Number     — the Point of Interconnection bus", False),
            ("    TS Fault Bus Number — the bus where TS faults will be applied", False),
            ("", False),
            ("Step 3:", False),
            ("  Fill From Bus No and To Bus No in TS_Contingencies sheet", False),
            ("  Fill From Bus No and To Bus No in PV_Contingencies sheet", False),
            ("  Fill Bus No in SC_Substations sheet", False),
            ("", False),
            ("Useful sheets in this file:", False),
            ("  All_Buses               — every bus in the model", False),
            ("  Generators              — type 2 (generator) and type 3 (slack) buses", False),
            ("  High_Voltage_138kV_plus — 138 kV and above (TS and SC monitoring)", False),
            ("  Buses_240kV             — 240 kV buses (connection voltage)", False),
            ("  Substation_Names        — one representative bus per substation name", False),
        ]

        for row_idx, (text, is_title) in enumerate(instructions, start=2):
            c = ws_inst.cell(row=row_idx, column=2, value=text)
            c.font = Font(
                name="Arial",
                size=11 if is_title else 10,
                bold=is_title,
                color="003865" if is_title else "333333",
            )

        wb.save(path)
        logger.info("Bus listing Excel saved: %s", path)

        # Print summary stats
        gen_count = len([b for b in self.buses if b.bus_type in (2, 3)])
        hv_count  = len([b for b in self.buses if b.base_kv >= 138.0])
        print(f"\n  Total buses : {len(self.buses)}")
        print(f"  Generators  : {gen_count}")
        print(f"  ≥138 kV     : {hv_count}")
        print(f"  Excel saved : {path}\n")

        return path

    def _extract_substation_names(self) -> list:
        """
        Extract one representative bus per substation name.
        PSS/E bus names often end with digits or 'S' (e.g. 'LANFINE 959S').
        Groups buses by their name prefix and picks the highest-kV bus.
        Returns list of (substation_label, BusEntry) sorted by base_kv desc.
        """
        seen: dict = {}
        for b in self.buses:
            name = b.bus_name.strip()
            if not name:
                continue
            # Use the full bus name as the substation label — most informative
            # for users matching to Study Scope substation names
            if name not in seen or b.base_kv > seen[name].base_kv:
                seen[name] = b

        result = sorted(seen.items(), key=lambda x: -x[1].base_kv)
        return result


# ── Entry point ───────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="bus_listing.py",
        description="List all buses from a PSS/E .sav file.",
    )
    parser.add_argument(
        "--sav",
        metavar="PATH",
        help="Path to the .sav file. Default: reads from config/settings.py.",
    )
    parser.add_argument(
        "--output",
        metavar="DIR",
        help="Output directory for Excel file. Default: reads from settings.py.",
    )
    parser.add_argument(
        "--filter",
        metavar="TEXT",
        help="Show only buses whose name contains TEXT (case-insensitive).",
    )
    parser.add_argument(
        "--kv",
        type=float,
        metavar="VOLTAGE",
        help="Show only buses at this base kV level (e.g. 240).",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args   = parser.parse_args()

    # Import settings
    from config.settings import PSSE_PATH, PSSE_VERSION, RESULTS_DIR

    sav_path   = args.sav    or USER_SAV_FILE
    output_dir = args.output or RESULTS_DIR

    logger.info("PSS/E path  : %s", PSSE_PATH)
    logger.info("SAV file    : %s", sav_path)
    logger.info("Output dir  : %s", output_dir)

    if not os.path.isfile(sav_path):
        logger.error(
            "SAV file not found: %s\n"
            "  Place your .sav file in the cases/ folder or use --sav flag.",
            sav_path
        )
        return 1

    listing = BusListing(psse_path=PSSE_PATH, psse_version=PSSE_VERSION)
    try:
        excel_path = listing.run(
            sav_path    = sav_path,
            output_dir  = output_dir,
            name_filter = args.filter,
            kv_filter   = args.kv,
        )
    except (RuntimeError, FileNotFoundError) as exc:
        logger.error("%s", exc)
        return 1

    if excel_path:
        print(f"  Next step: open {excel_path}")
        print("  Copy bus numbers into study_scope_data.xlsx Bus_Numbers sheet.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

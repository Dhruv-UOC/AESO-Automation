"""
project_io/excel_reader.py
---------------------------
Reads a filled study_scope_data.xlsx file and returns a ProjectData object.

All ten sheets are read in sequence. Missing sheets or empty rows are handled
gracefully — warnings are logged but reading continues. The caller receives a
fully populated ProjectData (with None values where data was not provided)
and can call project_data.validate() to get a list of warnings before running.

Usage
-----
    from project_io.excel_reader import ExcelReader

    reader = ExcelReader(r"D:\\Final_Project\\files\\projects\\P2611\\study_scope_data.xlsx")
    project = reader.read()
    warnings = project.validate()
    for w in warnings:
        print(w)
"""

import logging
import os
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from project_io.project_data import (
    BusNumberEntry,
    GeneratorDispatch,
    IntertieFow,
    ProjectData,
    ProjectInfo,
    PVContingency,
    RenewableDispatch,
    SCSubstation,
    Scenario,
    StudyMatrixEntry,
    TSContingency,
)

logger = logging.getLogger(__name__)

# Expected sheet names — must match the template exactly
SHEET_PROJECT_INFO    = "Project_Info"
SHEET_SCENARIOS       = "Scenarios"
SHEET_STUDY_MATRIX    = "Study_Matrix"
SHEET_CONV_GEN        = "Conv_Generation"
SHEET_RENEWABLES      = "Renewable_Dispatch"
SHEET_INTERTIE        = "Intertie_Flows"
SHEET_TS_CONT         = "TS_Contingencies"
SHEET_SC_SUBS         = "SC_Substations"
SHEET_PV_CONT         = "PV_Contingencies"
SHEET_BUS_NUMBERS     = "Bus_Numbers"

ALL_SHEETS = [
    SHEET_PROJECT_INFO,
    SHEET_SCENARIOS,
    SHEET_STUDY_MATRIX,
    SHEET_CONV_GEN,
    SHEET_RENEWABLES,
    SHEET_INTERTIE,
    SHEET_TS_CONT,
    SHEET_SC_SUBS,
    SHEET_PV_CONT,
    SHEET_BUS_NUMBERS,
]


class ExcelReaderError(Exception):
    """Raised for unrecoverable errors during Excel reading."""


class ExcelReader:
    """
    Reads study_scope_data.xlsx → ProjectData.

    Parameters
    ----------
    xlsx_path : str
        Absolute path to the filled study_scope_data.xlsx file.
    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self._wb: Optional[openpyxl.Workbook] = None

    # ── Public ────────────────────────────────────────────────────────────────

    def read(self) -> ProjectData:
        """
        Read all sheets and return a populated ProjectData.

        Raises ExcelReaderError if the file cannot be opened.
        """
        if not os.path.isfile(self.xlsx_path):
            raise ExcelReaderError(
                f"Excel file not found: {self.xlsx_path}"
            )

        logger.info("Reading study scope data: %s", self.xlsx_path)

        try:
            self._wb = openpyxl.load_workbook(
                self.xlsx_path,
                read_only=False,   # need write access for GUI saves
                data_only=True,    # read cell values, not formulas
            )
        except Exception as exc:
            raise ExcelReaderError(
                f"Could not open Excel file: {self.xlsx_path}\n"
                f"Reason: {exc}"
            ) from exc

        # Check which sheets are present
        present = set(self._wb.sheetnames)
        for sheet in ALL_SHEETS:
            if sheet not in present:
                logger.warning(
                    "Sheet '%s' not found in Excel file. "
                    "Related data will be empty.", sheet
                )

        project = ProjectData()

        project.info              = self._read_project_info()
        project.scenarios         = self._read_scenarios()
        project.study_matrix      = self._read_study_matrix()
        project.conv_gen          = self._read_conv_generation()
        project.renewables        = self._read_renewable_dispatch()
        project.intertie_flows    = self._read_intertie_flows()
        project.ts_contingencies  = self._read_ts_contingencies()
        project.sc_substations    = self._read_sc_substations()
        project.pv_contingencies  = self._read_pv_contingencies()
        project.bus_numbers       = self._read_bus_numbers()

        # Set project_dir and output_dir from xlsx location
        project.project_dir = os.path.dirname(self.xlsx_path)
        project.output_dir  = os.path.join(project.project_dir, "output")

        logger.info("Loaded: %s", str(project))
        return project

    # ── Private: sheet readers ────────────────────────────────────────────────

    def _read_project_info(self) -> ProjectInfo:
        """
        Sheet: Project_Info
        Layout: two columns — Field (col A) and Value (col B).
        Rows are matched by field name (case-insensitive, spaces/underscores ignored).
        """
        ws = self._get_sheet(SHEET_PROJECT_INFO)
        if ws is None:
            return ProjectInfo()

        # Build a dict: normalised_key → raw_value
        kv: Dict[str, Any] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 2:
                continue
            key = row[0]
            val = row[1]
            if key is None:
                continue
            norm = str(key).strip().lower().replace(" ", "_").replace("-", "_")
            kv[norm] = val

        info = ProjectInfo()
        info.project_number        = _str(kv.get("project_number"))
        info.project_name          = _str(kv.get("project_name"))
        info.market_participant    = _str(kv.get("market_participant"))
        info.studies_consultant    = _str(kv.get("studies_consultant"))
        info.in_service_date       = _str(kv.get("in_service_date"))
        info.generation_type       = _str(kv.get("generation_type", "Solar"))
        info.marp_mw               = _float(kv.get("marp_(mw)") or kv.get("marp_mw"))
        info.max_capability_mw     = _float(kv.get("maximum_capability_(mw)")
                                            or kv.get("max_capability_mw"))
        info.requested_sts_mw      = _float(kv.get("requested_rate_sts_(mw)")
                                            or kv.get("requested_sts_mw"))
        info.connection_voltage_kv = _float(kv.get("connection_voltage_(kv)")
                                            or kv.get("connection_voltage_kv"), 240.0)
        info.poc_substation_name   = _str(kv.get("poc_substation_name"))
        info.study_area_regions    = _str(kv.get("study_area_planning_areas")
                                         or kv.get("study_area_regions"))
        info.sav_file_path         = _str(kv.get("sav_file_path"))
        info.source_bus_number     = _int_or_none(kv.get("source_bus_number"))
        info.poi_bus_number        = _int_or_none(kv.get("poi_bus_number"))
        info.ts_fault_bus_number   = _int_or_none(kv.get("ts_fault_bus_number"))
        info.machine_id            = _str(kv.get("machine_id", "1")) or "1"

        logger.info(
            "Project: %s — %s  MARP=%.0f MW",
            info.project_number, info.project_name, info.marp_mw
        )
        return info

    def _read_scenarios(self) -> List[Scenario]:
        """
        Sheet: Scenarios
        Fixed columns:
        A: Scenario No | B: Year | C: Season | D: Dispatch Condition |
        E: Scenario Name | F: Pre/Post | G: Project Load (MW) | H: Project Gen (MW)
        """
        ws = self._get_sheet(SHEET_SCENARIOS)
        if ws is None:
            return []

        scenarios = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                sc = Scenario(
                    scenario_no     = _int(row[0] if len(row) > 0 else None, 0),
                    year            = _int(row[1] if len(row) > 1 else None, 0),
                    season          = _str(row[2] if len(row) > 2 else None),
                    dispatch_cond   = _str(row[3] if len(row) > 3 else None),
                    scenario_name   = _str(row[4] if len(row) > 4 else None),
                    pre_post        = _str(row[5] if len(row) > 5 else None),
                    project_load_mw = _float(row[6] if len(row) > 6 else None),
                    project_gen_mw  = _float(row[7] if len(row) > 7 else None),
                )
                if sc.scenario_name:
                    scenarios.append(sc)
            except Exception as exc:
                logger.warning("Skipping Scenarios row %s: %s", row, exc)

        logger.info("Loaded %d scenarios", len(scenarios))
        return scenarios

    def _read_study_matrix(self) -> List[StudyMatrixEntry]:
        """
        Sheet: Study_Matrix
        Row 1: headers
        Columns:
        A: Scenario Name
        B: Power Flow Cat A | C: Power Flow Cat B
        D: Volt Stability Cat A | E: Volt Stability Cat B
        F: Transient Cat A | G: Transient Cat B | H: Transient Conditional (X*)
        I: Motor Starting Cat A | J: Motor Starting Cat B
        K: Short Circuit Cat A

        Cells contain "X", "X*", or blank.
        """
        ws = self._get_sheet(SHEET_STUDY_MATRIX)
        if ws is None:
            return []

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                def _x(val) -> bool:
                    """True if cell contains X or X* (case-insensitive)."""
                    return str(val).strip().upper().startswith("X") if val else False

                entry = StudyMatrixEntry(
                    scenario_name        = _str(row[0] if len(row) > 0 else None),
                    power_flow_cat_a     = _x(row[1] if len(row) > 1 else None),
                    power_flow_cat_b     = _x(row[2] if len(row) > 2 else None),
                    volt_stability_cat_a = _x(row[3] if len(row) > 3 else None),
                    volt_stability_cat_b = _x(row[4] if len(row) > 4 else None),
                    transient_cat_a      = _x(row[5] if len(row) > 5 else None),
                    transient_cat_b      = _x(row[6] if len(row) > 6 else None),
                    transient_conditional= (
                        str(row[7]).strip().upper() == "X*"
                        if len(row) > 7 and row[7] else False
                    ),
                    motor_starting_cat_a = _x(row[8]  if len(row) > 8  else None),
                    motor_starting_cat_b = _x(row[9]  if len(row) > 9  else None),
                    short_circuit_cat_a  = _x(row[10] if len(row) > 10 else None),
                )
                if entry.scenario_name:
                    entries.append(entry)
            except Exception as exc:
                logger.warning("Skipping Study_Matrix row %s: %s", row, exc)

        logger.info("Loaded %d study matrix entries", len(entries))
        return entries

    def _read_conv_generation(self) -> List[GeneratorDispatch]:
        """
        Sheet: Conv_Generation
        Fixed columns A–E: Facility Name, Unit No, Bus No, MC (MW), Area No
        Then pairs of columns: Season Name, Season MW (repeating, any count)
        e.g. F="2028 SP", G=31.0, H="2028 SL", I=31.0, J="2028 WP", K=31.0
        """
        ws = self._get_sheet(SHEET_CONV_GEN)
        if ws is None:
            return []

        # Read header row to find season column pairs
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        season_cols = self._find_season_column_pairs(header, fixed_cols=5)

        generators = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                dispatch = {}
                for season_name, mw_col in season_cols:
                    mw = _float(row[mw_col] if mw_col < len(row) else None)
                    if season_name:
                        dispatch[season_name] = mw

                gen = GeneratorDispatch(
                    facility_name = _str(row[0] if len(row) > 0 else None),
                    unit_no       = _str(row[1] if len(row) > 1 else None),
                    bus_no        = _int_or_none(row[2] if len(row) > 2 else None),
                    mc_mw         = _float(row[3] if len(row) > 3 else None),
                    area_no       = _int(row[4] if len(row) > 4 else None, 0),
                    dispatch_mw   = dispatch,
                )
                if gen.facility_name:
                    generators.append(gen)
            except Exception as exc:
                logger.warning("Skipping Conv_Generation row %s: %s", row, exc)

        logger.info("Loaded %d conventional generators", len(generators))
        return generators

    def _read_renewable_dispatch(self) -> List[RenewableDispatch]:
        """
        Sheet: Renewable_Dispatch
        Fixed columns A–F: Facility Name, Type, Bus No, MC (MW), Area No
        Then pairs: Season Name, Season MW (same pattern as Conv_Generation)
        """
        ws = self._get_sheet(SHEET_RENEWABLES)
        if ws is None:
            return []

        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        season_cols = self._find_season_column_pairs(header, fixed_cols=5)

        renewables = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                dispatch = {}
                for season_name, mw_col in season_cols:
                    mw = _float(row[mw_col] if mw_col < len(row) else None)
                    if season_name:
                        dispatch[season_name] = mw

                ren = RenewableDispatch(
                    facility_name = _str(row[0] if len(row) > 0 else None),
                    gen_type      = _str(row[1] if len(row) > 1 else None),
                    bus_no        = _int_or_none(row[2] if len(row) > 2 else None),
                    mc_mw         = _float(row[3] if len(row) > 3 else None),
                    area_no       = _int(row[4] if len(row) > 4 else None, 0),
                    dispatch_mw   = dispatch,
                )
                if ren.facility_name:
                    renewables.append(ren)
            except Exception as exc:
                logger.warning("Skipping Renewable_Dispatch row %s: %s", row, exc)

        logger.info("Loaded %d renewable generators", len(renewables))
        return renewables

    def _read_intertie_flows(self) -> List[IntertieFow]:
        """
        Sheet: Intertie_Flows
        Column A: Scenario Name
        Columns B onward: one column per intertie (header = intertie name)
        e.g. B="AB-BC (MW)", C="AB-SK (MW)", D="MATL (MW)", E="City of Medicine Hat (MW)"
        Number of intertie columns varies by project — reader handles any count.
        """
        ws = self._get_sheet(SHEET_INTERTIE)
        if ws is None:
            return []

        # Read header to get intertie names from col B onward
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        intertie_names = []
        for i, h in enumerate(header[1:], start=1):
            if h:
                # Strip "(MW)" suffix if present for clean key
                name = str(h).replace("(MW)", "").replace("(mw)", "").strip()
                intertie_names.append((i, name))

        flows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                scenario_name = _str(row[0] if len(row) > 0 else None)
                if not scenario_name:
                    continue
                flow_dict = {}
                for col_idx, name in intertie_names:
                    val = _float(row[col_idx] if col_idx < len(row) else None)
                    flow_dict[name] = val
                flows.append(IntertieFow(
                    scenario_name=scenario_name,
                    flows=flow_dict,
                ))
            except Exception as exc:
                logger.warning("Skipping Intertie_Flows row %s: %s", row, exc)

        logger.info("Loaded intertie flows for %d scenarios", len(flows))
        return flows

    def _read_ts_contingencies(self) -> List[TSContingency]:
        """
        Sheet: TS_Contingencies
        Columns:
        A: Contingency Name | B: From Bus Name | C: To Bus Name |
        D: From Bus No | E: To Bus No | F: Circuit ID |
        G: Fault Location | H: Near End (cycles) | I: Far End (cycles)
        """
        ws = self._get_sheet(SHEET_TS_CONT)
        if ws is None:
            return []

        contingencies = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                cont = TSContingency(
                    contingency_name  = _str(row[0] if len(row) > 0 else None),
                    from_bus_name     = _str(row[1] if len(row) > 1 else None),
                    to_bus_name       = _str(row[2] if len(row) > 2 else None),
                    from_bus_no       = _int_or_none(row[3] if len(row) > 3 else None),
                    to_bus_no         = _int_or_none(row[4] if len(row) > 4 else None),
                    circuit_id        = _str(row[5] if len(row) > 5 else None) or "1",
                    fault_location    = _str(row[6] if len(row) > 6 else None),
                    near_end_cycles   = _float(row[7] if len(row) > 7 else None, 6.0),
                    far_end_cycles    = _float(row[8] if len(row) > 8 else None, 8.0),
                )
                if cont.contingency_name:
                    contingencies.append(cont)
            except Exception as exc:
                logger.warning("Skipping TS_Contingencies row %s: %s", row, exc)

        logger.info("Loaded %d TS contingencies", len(contingencies))
        return contingencies

    def _read_sc_substations(self) -> List[SCSubstation]:
        """
        Sheet: SC_Substations
        Columns: A: Substation Name | B: Bus No | C: Notes
        """
        ws = self._get_sheet(SHEET_SC_SUBS)
        if ws is None:
            return []

        substations = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                sub = SCSubstation(
                    substation_name = _str(row[0] if len(row) > 0 else None),
                    bus_no          = _int_or_none(row[1] if len(row) > 1 else None),
                    notes           = _str(row[2] if len(row) > 2 else None),
                )
                if sub.substation_name:
                    substations.append(sub)
            except Exception as exc:
                logger.warning("Skipping SC_Substations row %s: %s", row, exc)

        logger.info("Loaded %d SC target substations", len(substations))
        return substations

    def _read_pv_contingencies(self) -> List[PVContingency]:
        """
        Sheet: PV_Contingencies
        Columns:
        A: Contingency Name | B: From Bus Name | C: To Bus Name |
        D: From Bus No | E: To Bus No | F: Circuit ID | G: Category
        """
        ws = self._get_sheet(SHEET_PV_CONT)
        if ws is None:
            return []

        contingencies = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                cont = PVContingency(
                    contingency_name = _str(row[0] if len(row) > 0 else None),
                    from_bus_name    = _str(row[1] if len(row) > 1 else None),
                    to_bus_name      = _str(row[2] if len(row) > 2 else None),
                    from_bus_no      = _int_or_none(row[3] if len(row) > 3 else None),
                    to_bus_no        = _int_or_none(row[4] if len(row) > 4 else None),
                    circuit_id       = _str(row[5] if len(row) > 5 else None) or "1",
                    category         = _str(row[6] if len(row) > 6 else None) or "B",
                )
                if cont.contingency_name:
                    contingencies.append(cont)
            except Exception as exc:
                logger.warning("Skipping PV_Contingencies row %s: %s", row, exc)

        logger.info("Loaded %d PV contingencies", len(contingencies))
        return contingencies

    def _read_bus_numbers(self) -> List[BusNumberEntry]:
        """
        Sheet: Bus_Numbers
        Columns:
        A: Substation Name | B: Bus Number | C: Base kV | D: Bus Type | E: Notes
        """
        ws = self._get_sheet(SHEET_BUS_NUMBERS)
        if ws is None:
            return []

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            try:
                entry = BusNumberEntry(
                    substation_name = _str(row[0] if len(row) > 0 else None),
                    bus_number      = _int_or_none(row[1] if len(row) > 1 else None),
                    base_kv         = _float(row[2] if len(row) > 2 else None),
                    bus_type        = _int(row[3] if len(row) > 3 else None, 1),
                    notes           = _str(row[4] if len(row) > 4 else None),
                )
                if entry.substation_name:
                    entries.append(entry)
            except Exception as exc:
                logger.warning("Skipping Bus_Numbers row %s: %s", row, exc)

        logger.info("Loaded %d bus number entries", len(entries))
        return entries

    # ── Private: helpers ──────────────────────────────────────────────────────

    def _get_sheet(self, name: str) -> Optional[Worksheet]:
        """Return worksheet by name, or None with a warning if not present."""
        if self._wb is None:
            return None
        if name not in self._wb.sheetnames:
            logger.warning("Sheet '%s' not found — skipping.", name)
            return None
        return self._wb[name]

    @staticmethod
    def _find_season_column_pairs(
        header: List[Any],
        fixed_cols: int,
    ) -> List[tuple]:
        """
        Find (season_name, mw_col_index) pairs from a header row.

        After the fixed columns, the template uses alternating pairs:
        [Season Name col] [Season MW col] [Season Name col] [Season MW col] ...

        The season name cell contains the label (e.g. "2028 SP").
        The MW cell contains the value for that season.

        Returns list of (season_label, mw_column_index) tuples.
        """
        pairs = []
        i = fixed_cols
        while i < len(header) - 1:
            season_label = header[i]
            if season_label and str(season_label).strip():
                mw_col = i + 1
                pairs.append((str(season_label).strip(), mw_col))
                i += 2
            else:
                i += 1
        return pairs


# ─────────────────────────────────────────────────────────────────────────────
# Module-level type coercion helpers
# These handle None, empty string, and wrong types gracefully.
# ─────────────────────────────────────────────────────────────────────────────

def _str(val: Any, default: str = "") -> str:
    """Coerce to str, strip whitespace. Returns default if None or empty."""
    if val is None:
        return default
    result = str(val).strip()
    return result if result else default


def _float(val: Any, default: float = 0.0) -> float:
    """Coerce to float. Returns default on failure."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int(val: Any, default: int = 0) -> int:
    """Coerce to int. Returns default on failure."""
    if val is None:
        return default
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _int_or_none(val: Any) -> Optional[int]:
    """Coerce to int, return None if val is None, empty, or not numeric."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "n/a", "-", ""):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _is_empty_row(row: tuple) -> bool:
    """Return True if all cells in the row are None or empty strings."""
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in row
    )

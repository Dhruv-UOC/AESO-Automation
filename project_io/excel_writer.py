"""
project_io/excel_writer.py
---------------------------
Writes a ProjectData object back to study_scope_data.xlsx.

Used by the GUI when the user edits data in the editable tabs and
clicks Save. Also used to create a new project Excel file by copying
the blank template and immediately writing default values.

The writer updates ONLY the data cells (rows 2 onward in each sheet).
It preserves all formatting, headers, and column widths from the
template — it never recreates the workbook from scratch.

Usage
-----
    from project_io.excel_writer import ExcelWriter
    from project_io.project_data import ProjectData

    writer = ExcelWriter(r"D:\\...\\study_scope_data.xlsx")
    writer.write(project_data)
"""

import logging
import os
import shutil
from typing import Any, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from project_io.project_data import ProjectData

logger = logging.getLogger(__name__)

# Sheet names — must match template exactly
SHEET_PROJECT_INFO = "Project_Info"
SHEET_SCENARIOS    = "Scenarios"
SHEET_STUDY_MATRIX = "Study_Matrix"
SHEET_CONV_GEN     = "Conv_Generation"
SHEET_RENEWABLES   = "Renewable_Dispatch"
SHEET_INTERTIE     = "Intertie_Flows"
SHEET_TS_CONT      = "TS_Contingencies"
SHEET_SC_SUBS      = "SC_Substations"
SHEET_PV_CONT      = "PV_Contingencies"
SHEET_BUS_NUMBERS  = "Bus_Numbers"


class ExcelWriter:
    """
    Writes ProjectData back to an existing study_scope_data.xlsx file.

    Parameters
    ----------
    xlsx_path : str
        Path to the Excel file to update. Must already exist
        (created from the template by new_project_file()).
    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    # ── Public ────────────────────────────────────────────────────────────────

    def write(self, project: ProjectData) -> None:
        """
        Write all ProjectData fields back to the Excel file.
        Existing formatting is preserved — only cell values change.
        """
        if not os.path.isfile(self.xlsx_path):
            raise FileNotFoundError(
                f"Excel file not found: {self.xlsx_path}\n"
                f"Create it first using ExcelWriter.new_project_file()."
            )

        wb = openpyxl.load_workbook(self.xlsx_path, data_only=False)

        self._write_project_info(wb, project)
        self._write_scenarios(wb, project)
        self._write_study_matrix(wb, project)
        self._write_conv_generation(wb, project)
        self._write_renewable_dispatch(wb, project)
        self._write_intertie_flows(wb, project)
        self._write_ts_contingencies(wb, project)
        self._write_sc_substations(wb, project)
        self._write_pv_contingencies(wb, project)
        self._write_bus_numbers(wb, project)

        wb.save(self.xlsx_path)
        logger.info("Saved ProjectData to: %s", self.xlsx_path)

    @staticmethod
    def new_project_file(
        template_path: str,
        destination_path: str,
        overwrite: bool = False,
    ) -> None:
        """
        Create a new study_scope_data.xlsx for a project by copying
        the blank template.

        Parameters
        ----------
        template_path : str
            Path to templates/study_scope_template.xlsx.
        destination_path : str
            Where to write the new file.
        overwrite : bool
            If False (default), raises FileExistsError if destination exists.
        """
        if not os.path.isfile(template_path):
            raise FileNotFoundError(
                f"Template not found: {template_path}\n"
                f"Ensure templates/study_scope_template.xlsx exists."
            )

        if os.path.isfile(destination_path) and not overwrite:
            raise FileExistsError(
                f"File already exists: {destination_path}\n"
                f"Pass overwrite=True to replace it."
            )

        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        shutil.copy2(template_path, destination_path)
        logger.info("Created new project file: %s", destination_path)

    # ── Private: sheet writers ────────────────────────────────────────────────

    def _write_project_info(self, wb, project: ProjectData) -> None:
        """
        Sheet: Project_Info
        Layout: col A = field name (read-only), col B = value (we write this).
        Rows start at 2.
        """
        ws = self._get_sheet(wb, SHEET_PROJECT_INFO)
        if ws is None:
            return

        info = project.info
        # Map normalised field names to values
        values = {
            "project_number":              info.project_number,
            "project_name":                info.project_name,
            "market_participant":           info.market_participant,
            "studies_consultant":           info.studies_consultant,
            "in_service_date":             info.in_service_date,
            "generation_type":             info.generation_type,
            "marp_(mw)":                   info.marp_mw or "",
            "maximum_capability_(mw)":     info.max_capability_mw or "",
            "requested_rate_sts_(mw)":     info.requested_sts_mw or "",
            "connection_voltage_(kv)":     info.connection_voltage_kv or "",
            "poc_substation_name":         info.poc_substation_name,
            "study_area_planning_areas":   info.study_area_regions,
            "sav_file_path":               info.sav_file_path,
            "source_bus_number":           info.source_bus_number or "",
            "poi_bus_number":              info.poi_bus_number or "",
            "ts_fault_bus_number":         info.ts_fault_bus_number or "",
            "machine_id":                  info.machine_id,
        }

        for row in ws.iter_rows(min_row=2, max_col=2):
            key_cell = row[0]
            val_cell = row[1]
            if key_cell.value is None:
                continue
            norm = (str(key_cell.value).strip().lower()
                    .replace(" ", "_").replace("-", "_"))
            if norm in values:
                val_cell.value = values[norm]

    def _write_scenarios(self, wb, project: ProjectData) -> None:
        """Sheet: Scenarios — rows from row 3 onward (row 2 = header)."""
        ws = self._get_sheet(wb, SHEET_SCENARIOS)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=3, ncols=8)
        for i, sc in enumerate(project.scenarios):
            r = 3 + i
            _set(ws, r, 1, sc.scenario_no)
            _set(ws, r, 2, sc.year)
            _set(ws, r, 3, sc.season)
            _set(ws, r, 4, sc.dispatch_cond)
            _set(ws, r, 5, sc.scenario_name)
            _set(ws, r, 6, sc.pre_post)
            _set(ws, r, 7, sc.project_load_mw)
            _set(ws, r, 8, sc.project_gen_mw)

    def _write_study_matrix(self, wb, project: ProjectData) -> None:
        """Sheet: Study_Matrix."""
        ws = self._get_sheet(wb, SHEET_STUDY_MATRIX)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=3, ncols=11)

        def _x(flag: bool, conditional: bool = False) -> str:
            if conditional:
                return "X*"
            return "X" if flag else ""

        for i, entry in enumerate(project.study_matrix):
            r = 3 + i
            _set(ws, r, 1,  entry.scenario_name)
            _set(ws, r, 2,  _x(entry.power_flow_cat_a))
            _set(ws, r, 3,  _x(entry.power_flow_cat_b))
            _set(ws, r, 4,  _x(entry.volt_stability_cat_a))
            _set(ws, r, 5,  _x(entry.volt_stability_cat_b))
            _set(ws, r, 6,  _x(entry.transient_cat_a))
            _set(ws, r, 7,  _x(entry.transient_cat_b))
            _set(ws, r, 8,  _x(False, entry.transient_conditional))
            _set(ws, r, 9,  _x(entry.motor_starting_cat_a))
            _set(ws, r, 10, _x(entry.motor_starting_cat_b))
            _set(ws, r, 11, _x(entry.short_circuit_cat_a))

    def _write_conv_generation(self, wb, project: ProjectData) -> None:
        """
        Sheet: Conv_Generation
        Fixed cols 1-5, then alternating (season_name, season_mw) pairs.
        Season names come from the existing header row (row 2).
        """
        ws = self._get_sheet(wb, SHEET_CONV_GEN)
        if ws is None:
            return

        season_cols = self._read_season_header(ws, fixed_cols=5)
        self._clear_data_rows(ws, start_row=3, ncols=5 + len(season_cols) * 2)

        for i, gen in enumerate(project.conv_gen):
            r = 3 + i
            _set(ws, r, 1, gen.facility_name)
            _set(ws, r, 2, gen.unit_no)
            _set(ws, r, 3, gen.bus_no if gen.bus_no is not None else "")
            _set(ws, r, 4, gen.mc_mw)
            _set(ws, r, 5, gen.area_no)
            for season_name, name_col, mw_col in season_cols:
                _set(ws, r, name_col, season_name)
                mw = gen.dispatch_mw.get(season_name, "")
                _set(ws, r, mw_col, mw)

    def _write_renewable_dispatch(self, wb, project: ProjectData) -> None:
        """Sheet: Renewable_Dispatch."""
        ws = self._get_sheet(wb, SHEET_RENEWABLES)
        if ws is None:
            return

        season_cols = self._read_season_header(ws, fixed_cols=5)
        self._clear_data_rows(ws, start_row=3, ncols=5 + len(season_cols) * 2)

        for i, ren in enumerate(project.renewables):
            r = 3 + i
            _set(ws, r, 1, ren.facility_name)
            _set(ws, r, 2, ren.gen_type)
            _set(ws, r, 3, ren.bus_no if ren.bus_no is not None else "")
            _set(ws, r, 4, ren.mc_mw)
            _set(ws, r, 5, ren.area_no)
            for season_name, name_col, mw_col in season_cols:
                _set(ws, r, name_col, season_name)
                mw = ren.dispatch_mw.get(season_name, "")
                _set(ws, r, mw_col, mw)

    def _write_intertie_flows(self, wb, project: ProjectData) -> None:
        """
        Sheet: Intertie_Flows
        Intertie column names come from header row (row 2) cols B onward.
        """
        ws = self._get_sheet(wb, SHEET_INTERTIE)
        if ws is None:
            return

        # Read intertie column names from header
        header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        intertie_cols = []
        for i, h in enumerate(header[1:], start=2):
            if h:
                name = str(h).replace("(MW)", "").replace("(mw)", "").strip()
                intertie_cols.append((i, name))

        self._clear_data_rows(ws, start_row=3, ncols=len(header))

        for i, flow in enumerate(project.intertie_flows):
            r = 3 + i
            _set(ws, r, 1, flow.scenario_name)
            for col_idx, intertie_name in intertie_cols:
                val = flow.flows.get(intertie_name, "")
                _set(ws, r, col_idx, val)

    def _write_ts_contingencies(self, wb, project: ProjectData) -> None:
        """Sheet: TS_Contingencies."""
        ws = self._get_sheet(wb, SHEET_TS_CONT)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=3, ncols=9)
        for i, cont in enumerate(project.ts_contingencies):
            r = 3 + i
            _set(ws, r, 1, cont.contingency_name)
            _set(ws, r, 2, cont.from_bus_name)
            _set(ws, r, 3, cont.to_bus_name)
            _set(ws, r, 4, cont.from_bus_no if cont.from_bus_no is not None else "")
            _set(ws, r, 5, cont.to_bus_no   if cont.to_bus_no   is not None else "")
            _set(ws, r, 6, cont.circuit_id)
            _set(ws, r, 7, cont.fault_location)
            _set(ws, r, 8, cont.near_end_cycles)
            _set(ws, r, 9, cont.far_end_cycles)

    def _write_sc_substations(self, wb, project: ProjectData) -> None:
        """Sheet: SC_Substations."""
        ws = self._get_sheet(wb, SHEET_SC_SUBS)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=3, ncols=3)
        for i, sub in enumerate(project.sc_substations):
            r = 3 + i
            _set(ws, r, 1, sub.substation_name)
            _set(ws, r, 2, sub.bus_no if sub.bus_no is not None else "")
            _set(ws, r, 3, sub.notes)

    def _write_pv_contingencies(self, wb, project: ProjectData) -> None:
        """Sheet: PV_Contingencies."""
        ws = self._get_sheet(wb, SHEET_PV_CONT)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=3, ncols=7)
        for i, cont in enumerate(project.pv_contingencies):
            r = 3 + i
            _set(ws, r, 1, cont.contingency_name)
            _set(ws, r, 2, cont.from_bus_name)
            _set(ws, r, 3, cont.to_bus_name)
            _set(ws, r, 4, cont.from_bus_no if cont.from_bus_no is not None else "")
            _set(ws, r, 5, cont.to_bus_no   if cont.to_bus_no   is not None else "")
            _set(ws, r, 6, cont.circuit_id)
            _set(ws, r, 7, cont.category)

    def _write_bus_numbers(self, wb, project: ProjectData) -> None:
        """Sheet: Bus_Numbers."""
        ws = self._get_sheet(wb, SHEET_BUS_NUMBERS)
        if ws is None:
            return

        self._clear_data_rows(ws, start_row=4, ncols=5)
        for i, entry in enumerate(project.bus_numbers):
            r = 4 + i
            _set(ws, r, 1, entry.substation_name)
            _set(ws, r, 2, entry.bus_number if entry.bus_number is not None else "")
            _set(ws, r, 3, entry.base_kv)
            _set(ws, r, 4, entry.bus_type)
            _set(ws, r, 5, entry.notes)

    # ── Private: helpers ──────────────────────────────────────────────────────

    @staticmethod
    def _get_sheet(wb, name: str) -> Optional[Worksheet]:
        if name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook — skipping.", name)
            return None
        return wb[name]

    @staticmethod
    def _clear_data_rows(ws: Worksheet, start_row: int, ncols: int) -> None:
        """Clear cell values in data rows without removing formatting."""
        for row in ws.iter_rows(min_row=start_row, max_col=ncols):
            for cell in row:
                cell.value = None

    @staticmethod
    def _read_season_header(
        ws: Worksheet,
        fixed_cols: int,
    ) -> List[tuple]:
        """
        Read (season_name, name_col_idx, mw_col_idx) triples from the header.
        Header row = row 2 (row 1 is the note row in template).
        fixed_cols = number of fixed columns before season pairs start.
        """
        header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        pairs = []
        i = fixed_cols
        while i < len(header) - 1:
            season_label = header[i]
            if season_label and str(season_label).strip():
                # col indices in openpyxl are 1-based
                pairs.append((str(season_label).strip(), i + 1, i + 2))
                i += 2
            else:
                i += 1
        return pairs


# ─────────────────────────────────────────────────────────────────────────────
# Helper
# ─────────────────────────────────────────────────────────────────────────────

def _set(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """Set a cell value, leaving formatting intact."""
    ws.cell(row=row, column=col).value = value

"""
tests/test_golden.py
--------------------
Validates automation output against the golden_results.xlsx baseline.
Run with: pytest tests/test_golden.py -v
"""

import os
import sys
import pytest
import openpyxl

# ── Project root on path ───────────────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from config.settings import BASE_DIR, PROJECTS_DIR, PSSE_PATH, PSSE_VERSION, AESO

# ── Configure these for your project ──────────────────────────────────────────
PROJECT_NUMBER = None   # e.g. "P2611" or None for base case
# ──────────────────────────────────────────────────────────────────────────────

if PROJECT_NUMBER:
    RESULTS_DIR  = os.path.join(PROJECTS_DIR, PROJECT_NUMBER, "output", "results")
    SAV_FILE     = os.path.join(PROJECTS_DIR, PROJECT_NUMBER, "cases", "model.sav")
else:
    RESULTS_DIR  = os.path.join(BASE_DIR, "output")
    SAV_FILE     = os.path.join(BASE_DIR, "cases", "sample.sav")

GOLDEN_FILE = os.path.join(RESULTS_DIR, "golden_results.xlsx")

# ── Tolerances ─────────────────────────────────────────────────────────────────
TOL_VOLTAGE_PU  = 0.005   # ±0.5% on bus voltages
TOL_GEN_MW      = 1.0     # ±1 MW on total generation
TOL_LOAD_MW     = 1.0     # ±1 MW on total load
TOL_LOSS_MW     = 2.0     # ±2 MW on total losses


# ── Load golden Excel once at module level ─────────────────────────────────────
def _load_golden(path: str) -> dict:
    """Read golden_results.xlsx into a dict matching the old golden structure."""
    if not os.path.isfile(path):
        pytest.skip(f"Golden file not found — run extract_golden.py first.\n  Expected: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── System Totals sheet ────────────────────────────────────────────────────
    ws_totals = wb["System Totals"]
    system_totals = {}
    key_map = {
        "Generation": "generation_mw",
        "Load":       "load_mw",
        "Losses":     "losses_mw",
    }
    for row in ws_totals.iter_rows(min_row=2, values_only=True):
        if row[0] in key_map:
            system_totals[key_map[row[0]]] = float(row[1])

    # ── Bus Voltages sheet ─────────────────────────────────────────────────────
    ws_buses = wb["Bus Voltages"]
    buses = {}
    for row in ws_buses.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        buses[int(row[0])] = {
            "bus_number": int(row[0]),
            "bus_name":   str(row[1]).strip(),
            "base_kv":    float(row[2]),
            "voltage_pu": float(row[3]),
            "angle_deg":  float(row[4]),
            "bus_type":   int(row[5]),
        }

    # ── Branch Flows sheet ─────────────────────────────────────────────────────
    ws_branches = wb["Branch Flows"]
    branches = []
    for row in ws_branches.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        branches.append({
            "from_bus":    int(row[0]),
            "to_bus":      int(row[1]),
            "circuit_id":  str(row[2]).strip(),
            "mw_flow":     float(row[3]),
            "mvar_flow":   float(row[4]),
            "mva_flow":    float(row[5]),
            "rating_mva":  float(row[6]),
            "loading_pct": float(row[7]),
        })

    wb.close()
    return {"system_totals": system_totals, "buses": buses, "branches": branches}


GOLDEN = _load_golden(GOLDEN_FILE)


# ── Fixtures ───────────────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def pf_results():
    """Run the real power flow once and return results for all tests."""
    from core.psse_interface import PSSEInterface
    from studies.power_flow.power_flow_study import PowerFlowStudy

    psse = PSSEInterface(psse_path=PSSE_PATH, psse_version=PSSE_VERSION, mock=False)
    psse.initialize()
    study = PowerFlowStudy(
        psse,
        scenario_label="Golden_Validation",
        voltage_min=AESO["voltage_min_pu"],
        voltage_max=AESO["voltage_max_pu"],
        voltage_min_contingency=AESO["voltage_min_contingency"],
        voltage_max_contingency=AESO["voltage_max_contingency"],
        thermal_limit_pct=AESO["thermal_limit_pct"],
    )
    results = study.run(SAV_FILE)
    return results


# ── Tests ──────────────────────────────────────────────────────────────────────
def test_converged(pf_results):
    assert pf_results.converged, "Power flow did not converge"


def test_system_generation(pf_results):
    diff = abs(pf_results.total_generation_mw - GOLDEN["system_totals"]["generation_mw"])
    assert diff < TOL_GEN_MW, (
        f"Generation mismatch: got {pf_results.total_generation_mw:.2f} MW, "
        f"expected {GOLDEN['system_totals']['generation_mw']:.2f} MW, "
        f"diff={diff:.2f} MW"
    )


def test_system_load(pf_results):
    diff = abs(pf_results.total_load_mw - GOLDEN["system_totals"]["load_mw"])
    assert diff < TOL_LOAD_MW, (
        f"Load mismatch: got {pf_results.total_load_mw:.2f} MW, "
        f"expected {GOLDEN['system_totals']['load_mw']:.2f} MW, "
        f"diff={diff:.2f} MW"
    )


def test_system_losses(pf_results):
    diff = abs(pf_results.total_losses_mw - GOLDEN["system_totals"]["losses_mw"])
    assert diff < TOL_LOSS_MW, (
        f"Loss mismatch: got {pf_results.total_losses_mw:.2f} MW, "
        f"expected {GOLDEN['system_totals']['losses_mw']:.2f} MW, "
        f"diff={diff:.2f} MW"
    )


def test_key_bus_voltages(pf_results):
    """Check all buses that appear in both results and golden file."""
    results_bus_map = {b.bus_number: b for b in pf_results.buses}
    failures = []

    for bus_num, g_bus in GOLDEN["buses"].items():
        r_bus = results_bus_map.get(bus_num)
        if r_bus is None:
            continue
        diff = abs(r_bus.voltage_pu - g_bus["voltage_pu"])
        if diff >= TOL_VOLTAGE_PU:
            failures.append(
                f"  Bus {bus_num:>7} ({g_bus['bus_name']:<12}) "
                f"got={r_bus.voltage_pu:.5f}  "
                f"expected={g_bus['voltage_pu']:.5f}  "
                f"diff={diff:.5f} pu"
            )

    assert not failures, (
        f"{len(failures)} bus voltage(s) outside ±{TOL_VOLTAGE_PU} pu tolerance:\n"
        + "\n".join(failures[:20])   # show first 20 failures max
    )


def test_no_new_overloads(pf_results):
    """No branch should be loaded above the AESO thermal limit."""
    violations = [
        b for b in pf_results.branches
        if b.loading_pct > AESO["thermal_limit_pct"]
    ]
    assert not violations, (
        f"{len(violations)} thermal violation(s) found:\n"
        + "\n".join(
            f"  {v.from_bus}→{v.to_bus} [{v.circuit_id}]  {v.loading_pct:.1f}%"
            for v in violations[:10]
        )
    )